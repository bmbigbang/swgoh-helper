import json
from env import get_env
from initialise_data_structures import initialise_data_structures
from openpyxl import load_workbook
from data_lookups import mod_slots
from mod_optimizer import construct_weights, restructure_mods, score_and_sort_mods, column_stat_map, \
    col_to_mod_stat_map

data_lookups = initialise_data_structures()
env_data = get_env()

allycode = env_data["allycode"]

with open('saved-mods.json', 'r', encoding='utf-8') as f:
    saved_mods = json.load(f)

with open('./swgoh-stat-calc/unit_stat_data.json', 'r', encoding='utf-8') as f:
    unit_stats = json.load(f)

saved_data = initialise_data_structures()

workbook = load_workbook('mod-weights.xlsx')

human_readable_stat_names = {
    "Speed": "Speed",
    "SpeedPercentAdditive": "",
    "Offense": "Offense",
    "OffensePercentAdditive": "%Offense",
    "CriticalChancePercentAdditive": "Crit Chance",
    "CriticalDamage": "Crit Damage",
    "Health": "Health",
    "MaxHealthPercentAdditive": "%Heath",
    "Defense": "Defense",
    "DefensePercentAdditive": "%Defense",
    "MaxShield": "Prot",
    "MaxShieldPercentAdditive": "%Prot",
    "CriticalNegateChancePercentAdditive": "Crit Avoid",
    "Resistance": "Tenacity",
    "Accuracy": "Potency",
    "EvasionNegatePercentAdditive": "Accuracy"
}


def sort_mods():
    weights = construct_weights()
    recommendations = {}
    mod_map = {}
    for weight_object in weights:
        if weight_object["char_name"] not in saved_mods[allycode]["char_name_map"]:
            recommendations[weight_object["char_name"]] = None
            continue

        recommendations[weight_object["char_name"]] = []

        mods_container = restructure_mods(weight_object["char_name"], mod_map)
        slots = [i for i in mod_slots.values()]

        for slot in slots:
            recommendations[weight_object["char_name"]].append(
                score_and_sort_mods(mods_container, slot, weight_object["weights"], mod_map, [])
            )

    return recommendations, mod_map


def update_excel_with_recommendations():
    if "Sorts" in workbook.sheetnames:
        del workbook["Sorts"]
    worksheet = workbook.create_sheet("Sorts")
    worksheet.cell(row=1, column=1).value = "Characters (Sorted By Priority)"
    worksheet.cell(row=1, column=2).value = "Details"
    col = 3
    for idx, stat_names in col_to_mod_stat_map.items():
        for stat_name in stat_names:
            if stat_name == "SpeedPercentAdditive":
                continue
            worksheet.cell(row=1, column=col).value = human_readable_stat_names[stat_name]
            col += 1

    sorted_mods, mod_map = sort_mods()
    row = 2
    for char_name, sorted_mods_per_slot in sorted_mods.items():
        worksheet.cell(row=row, column=1).value = char_name

        if sorted_mods_per_slot is None:
            worksheet.cell(row=row, column=2).value = "Character is not lvl 51 or was excluded"
            row += 1
            continue
        elif len(sorted_mods_per_slot) == 0:
            worksheet.cell(row=row, column=2).value = "Ran out of mods"
            row += 1
            continue
        for x, slot in mod_slots.items():
            for y in range(len(sorted_mods_per_slot[x])):
                details = slot + ":\n"
                mod_id = mod_map[sorted_mods_per_slot[x][y][-1]]
                mod = saved_mods[allycode]["mods"][mod_id]
                details += "{} Primary - {} Set from {}\n".format(
                    mod["primary"],
                    mod["set"]["name"],
                    mod["char_name"]
                )
                worksheet.cell(row=row, column=2).value = details

                col = 3
                for idx, stat_names in col_to_mod_stat_map.items():
                    for stat_name in stat_names:
                        if stat_name == "SpeedPercentAdditive":
                            continue
                        if stat_name in mod["stats"]:
                            worksheet.cell(row=row, column=col).value = float(mod["stats"][stat_name])
                        col += 1

                row += 1

    workbook.save('mod-weights.xlsx')


update_excel_with_recommendations()