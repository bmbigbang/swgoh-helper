import json
import re
from env import get_env
from initialise_data_structures import initialise_data_structures
from openpyxl import load_workbook
from data_lookups import mod_slots, mod_set_stats
import numpy as np

data_lookups = initialise_data_structures()
env_data = get_env()

allycode = env_data["allycode"]

with open('saved-mods.json', 'r', encoding='utf-8') as f:
    saved_mods = json.load(f)

with open('./swgoh-stat-calc/unit_stat_data.json', 'r', encoding='utf-8') as f:
    unit_stats = json.load(f)

saved_data = initialise_data_structures()

workbook = load_workbook('mod-weights.xlsx')
char_mod_stat_weights = workbook["Char Mod Stat Weights"]

# this significantly impacts performance
candidates_to_try_in_each_slot = 8

column_stat_map = {
    1: "Speed",
    2: "Offence",
    3: "Crit Chance",
    4: "Crit Damage",
    5: "Health",
    6: "Defence",
    7: "Protection",
    8: "Crit Avoidance",
    9: "Tenacity",
    10: "Potency",
    11: "Accuracy"
}

col_to_mod_stat_map = {
    1: ["Speed", "SpeedPercentAdditive"],
    2: ["Offence", "OffensePercentAdditive"],
    3: ["CriticalChancePercentAdditive"],
    4: ["CriticalDamage"],
    5: ["Health", "MaxHealthPercentAdditive"],
    6: ["Defense", "DefensePercentAdditive"],
    7: ["MaxShield", "MaxShieldPercentAdditive"],
    8: ["CriticalNegateChancePercentAdditive"],
    9: ["Resistance"],
    10: ["Accuracy"],
    11: ["EvasionNegatePercentAdditive"]
}

# primary stats have an unfair advantage in this calc, by letting them normalise from 0
max_mod_values_per_stat = {
    "Speed": 32,
    "SpeedPercentAdditive": 0.1,
    "Offence": 253,
    "OffensePercentAdditive": 0.085,
    "CriticalChancePercentAdditive": 0.2,
    "CriticalDamage": 0.42,
    "Health": 2696,
    "MaxHealthPercentAdditive": 0.16,
    "Defense": 73,
    "DefensePercentAdditive": 0.2,
    "MaxShield": 4607,
    "MaxShieldPercentAdditive": 0.24,
    "CriticalNegateChancePercentAdditive": 0.35,
    "Resistance": 0.35,
    "Accuracy": 0.3,
    "EvasionNegatePercentAdditive": 0.3,
}

char_name_map = {j["nameKey"]: i for i, j in saved_data["toons"].items()}


def construct_weights():
    weights = []
    threshold_pattern = re.compile(r"(?P<init>\d+)(\s*,\s*)(?P<threshold>\d+\.?\d*)(\s*,\s*)(?P<final>\d+)")

    count = 2
    for row in char_mod_stat_weights.iter_rows(min_row=2, max_col=12, max_row=1000):

        char_name = row[0].value
        if not char_name:
            print("Empty char name on row {} of mod-weights.xlsx".format(count))
            break

        char_weights = []
        char_thresholds = []

        for col_idx, stat_name in column_stat_map.items():
            cell = row[col_idx].value
            if not cell:
                char_weights.append(0)
                char_thresholds.append({0: 0})
            elif type(cell) is int:
                char_weights.append(cell)
                char_thresholds.append({0: 0})
            elif cell.startswith("("):
                threshold = threshold_pattern.search(cell)
                if not threshold:
                    print("Could not validate threshold pattern on row {} - {}".format(count, stat_name))
                    char_weights.append(0)
                    char_thresholds.append({0: 0})
                else:
                    char_weights.append(int(threshold.groupdict()["init"]))
                    char_thresholds.append({
                        float(threshold.groupdict()["threshold"]): int(threshold.groupdict()["final"])
                    })

        char_thresholds.append({0: 0})
        char_weights.append(0)

        weights.append({
            "char_name": char_name,
            "weights": np.array(char_weights),
            "thresholds": char_thresholds
        })
        count += 1
    return weights


def find(pred, iterable):
    for element in iterable:
        if pred(element):
            return element
    return None


def scaled_stat_modifiers(stat_name, char_name):
    char_stats = unit_stats[char_name_map[char_name]]
    percent_modifier = max_mod_values_per_stat.get(stat_name + "PercentAdditive", 1)
    if stat_name == "Offense":
        return 1 / (char_stats["final"]["6"] * percent_modifier)
    elif stat_name == "Health":
        return 1 / (char_stats["final"]["1"] * percent_modifier)
    elif stat_name == "Defense":
        return 2 / ((char_stats["base"]["8"] + char_stats["base"]["9"]) * percent_modifier)
    elif stat_name == "MaxShield":
        if "28" not in char_stats["final"]:
            return 0
        return 1 / (char_stats["final"]["28"] * percent_modifier)
    elif stat_name == "SpeedPercentAdditive":
        return (0.1 * char_stats["final"]["5"]) / (max_mod_values_per_stat["Speed"] / max_mod_values_per_stat[stat_name])
    elif stat_name == "Speed":
        return 1 / max_mod_values_per_stat[stat_name]
    else:
        return 0.01 / max_mod_values_per_stat[stat_name]


def restructure_mods(char_name, mod_map):
    mods_container = {slot: [] for slot in mod_slots.values()}
    count = 2
    for mod_id, mod in saved_mods[allycode]["mods"].items():
        mod_vector = [0.0001 for _i in col_to_mod_stat_map.keys()]
        for col_idx, stat_names in col_to_mod_stat_map.items():
            for stat_name in stat_names:
                value = mod["stats"].get(stat_name, 0)
                if value:
                    scaled_value = scaled_stat_modifiers(stat_name, char_name)
                    mod_vector[col_idx - 1] += scaled_value * value

        mod_vector.append(count)
        nd_array = np.array(mod_vector)
        mods_container[mod["slot"]].append(nd_array)
        mod_map[count] = mod_id
        count += 1
    for slot in mod_slots.values():
        mods_container["np-" + slot] = np.array(mods_container[slot])

    mods_container["sets"] = np.zeros((len(mod_set_stats.keys()), len(col_to_mod_stat_map.keys()) + 1))
    for mod_set_idx, mod_set in mod_set_stats.items():
        for col_idx, stat_names in col_to_mod_stat_map.items():
            if mod_set["stat"][0] in stat_names:
                scaled_value = scaled_stat_modifiers(mod_set["stat"][0], char_name)
                mods_container["sets"][mod_set_idx - 1][col_idx - 1] += scaled_value * mod_set["stat"][1]
    return mods_container


def score_and_sort_mods(mods_container, slot, weights):
    scores_mods = np.inner(mods_container[slot], weights)
    indices = np.argpartition(scores_mods, -candidates_to_try_in_each_slot)[-candidates_to_try_in_each_slot:]
    candidates = mods_container["np-" + slot][indices]
    return candidates


def evaluate_sets(candidates, mod_map, mod_sets_container):
    sets = [saved_mods[allycode]["mods"][mod_map[i[-1]]]["set"] for i in candidates]
    set_names = [i["name"] for i in sets]
    set_stats = np.zeros(len(col_to_mod_stat_map.keys()) + 1)
    for mod_set_idx, mod_set in mod_set_stats.items():
        matching = len([i for i in set_names if i == mod_set["name"]]) / mod_set["setCount"]
        matching_sets = np.floor(matching)
        for _ in range(int(matching_sets)):
            set_stats += mod_sets_container[mod_set_idx - 1]
    return set_stats


def evaluate_thresholds(candidates, thresholds, weights, set_stats):
    score_sum = np.sum(candidates, axis=0) + set_stats
    for x in range(len(thresholds)):
        for threshold, new_weight in thresholds[x].items():
            if threshold and weights[x] != new_weight:
                if x == 0:
                    stat_upscaler = max_mod_values_per_stat[col_to_mod_stat_map[x + 1][0]]
                else:
                    stat_upscaler = max_mod_values_per_stat[col_to_mod_stat_map[x + 1][-1]]
                if score_sum[x] * stat_upscaler >= threshold:
                    weights[x] = new_weight

    return weights


def recurse_slots(weights, thresholds, candidates, slots, slot_id, mods_container, mod_map, score, char_name):
    slot_id += 1
    slot = slots[slot_id]
    sorted_mods = score_and_sort_mods(mods_container, slot, weights)
    set_stats = evaluate_sets(candidates, mod_map, mods_container["sets"])
    W = evaluate_thresholds(
        candidates,
        thresholds,
        np.copy(weights),
        set_stats
    )

    result_candidates = np.concatenate((candidates, np.array([sorted_mods[0]])))

    for candidate in sorted_mods:
        new_candidates = np.concatenate((candidates, np.array([candidate])))

        if len(slots) - 1 > slot_id and len(mods_container[slots[slot_id + 1]]):
            new_candidates, new_score = recurse_slots(
                W,
                thresholds,
                candidates=new_candidates,
                slots=slots,
                slot_id=slot_id,
                mods_container=mods_container,
                mod_map=mod_map,
                score=score,
                char_name=char_name
            )
        else:
        # print([saved_mods[allycode]["mods"][mod_map[str(i.data)]]["set"] for i in new_candidates])
            set_stats = evaluate_sets(candidates, mod_map, mods_container["sets"])
            new_score = np.sum(np.inner(np.concatenate((np.array([set_stats]), candidates)), weights))

        if new_score > score:
            score = new_score
            result_candidates = np.copy(new_candidates)

    return result_candidates, score


def evaluate_mods():
    weights = construct_weights()
    recommendations = {}
    mod_sets = {}
    used_mods = set()
    mod_map = {}
    for weight_object in weights[:6]:
        if char_name_map[weight_object["char_name"]] not in saved_mods[allycode]["chars"]:
            recommendations[weight_object["char_name"]] = None
            continue
        mods_container = restructure_mods(weight_object["char_name"], mod_map)
        slots = [i for i in mod_slots.values()]

        for slot in slots:
            ##  np.extract(condition, arr)
            mods_container[slot] = [i for i in filter(lambda x: x[-1] not in used_mods, mods_container[slot])]

        square_sorted_mods = score_and_sort_mods(mods_container, "Square", weight_object["weights"])
        mod_candidates = np.array([])
        score = np.sum(square_sorted_mods[0] * weight_object["weights"])

        for square_candidate in square_sorted_mods:
            new_candidates = np.array([square_candidate])

            if len(mods_container[slots[1]]):
                new_candidates, new_score = recurse_slots(
                    weight_object["weights"],
                    weight_object["thresholds"],
                    candidates=new_candidates,
                    slots=slots,
                    slot_id=0,
                    mods_container=mods_container,
                    mod_map=mod_map,
                    score=score,
                    char_name=weight_object["char_name"]
                )
            else:
                new_score = np.inner(new_candidates, weight_object["weights"])

            if new_score > score:
                mod_candidates = np.copy(new_candidates)
                score = new_score

        recommendations[weight_object["char_name"]] = mod_candidates
        used_mods.update(set([i[-1] for i in mod_candidates]))
        mod_sets[weight_object["char_name"]] = mods_container["sets"]

    return recommendations, mod_map, mod_sets


def update_excel_with_recommendations():
    slots = [i for i in mod_slots.values()]
    del workbook["Recommendations"]
    worksheet = workbook.create_sheet("Recommendations")
    worksheet.cell(row=1, column=1).value = "Characters (Sorted By Priority)"
    worksheet.cell(row=1, column=2).value = "Recommendations"
    for col_idx, stat_name in column_stat_map.items():
        worksheet.cell(row=1, column=col_idx + 2).value = stat_name

    recommendations, mod_map, mod_sets_container = evaluate_mods()
    count = 2
    for char_name, mod_candidates in recommendations.items():
        worksheet.cell(row=count, column=1).value = char_name

        if mod_candidates is None:
            print(char_name)
            worksheet.cell(row=count, column=2).value = "Character is not lvl 51 or was excluded"
            continue

        recommendation = ""
        for x in range(len(mod_candidates)):
            mod_id = mod_map[mod_candidates[x][-1]]
            mod = saved_mods[allycode]["mods"][mod_id]
            recommendation += "{} {} from {}{}".format(mod["set"]["name"], slots[x], mod["char_name"], "\n")
        worksheet.cell(row=count, column=2).value = recommendation

        set_stats = evaluate_sets(mod_candidates, mod_map, mod_sets_container[char_name])
        mods_vector = np.sum(mod_candidates, axis=0) + set_stats

        for col_idx, stat_names in col_to_mod_stat_map.items():
            stat_name = stat_names[-1]
            if col_idx == 1:
                stat_name = stat_names[0]

            worksheet.cell(row=count, column=col_idx + 2).value = float(mods_vector[col_idx - 1] * max_mod_values_per_stat[stat_name])

        count += 1
        # Some diagnostics here:
        # print("Recommendations for {}".format(char_name))
        #
        # for x in range(len(mod_candidates)):
        #     mod_id = mod_map[mod_candidates[x][-1]]
        #     mod = saved_mods[allycode]["mods"][mod_id]
        #     print(mod_id, "-", slots[x], "from {}".format(mod["char_name"]), " with {} set".format(mod["set"]["name"]))
        #
        # active_sets = evaluate_sets(mod_candidates, mod_map, char_name)
        # s = add_set_values(active_sets, mod_candidates)
        # stats = []
        # for x, stat_names in col_to_mod_stat_map.items():
        #     stat_name = stat_names[-1]
        #     if x == 1:
        #         stat_name = stat_names[0]
        #     stats.append(column_stat_map[x] + ": {:f}".format(float(s[x - 1] * max_mod_values_per_stat[stat_name])))
        # print("------------------------------------")
        # print("Total stats: ", ", ".join(stats))
        # print("------------------------------------")

    workbook.save('mod-weights.xlsx')


update_excel_with_recommendations()
